import os

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

PATH = 'daten'
RESULT = 'daten/result.xlsx'
FILTER = ['xls', 'xlsx']

result_workbook = Workbook(write_only=True)

for file in os.listdir(PATH):
    if file.rsplit('.')[1] in FILTER:
        file_path = os.path.join(PATH, file)
        workbook = load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            result_sheet = result_workbook.create_sheet(sheet_name)
            for row in sheet:
                result_sheet.append(row)

result_workbook.save(RESULT)
